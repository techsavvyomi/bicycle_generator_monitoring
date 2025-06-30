# tracker.py
import os
import time
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook

class SessionTracker:
    def __init__(self, cycle_id):
        self.cycle_id = cycle_id
        self.running = False
        self.start_time = None
        self.data = {}
        self.current_student = None
        self.total_voltage = 0

    def start(self, student):
        """Begin tracking a new session for `student`."""
        self.running = True
        self.start_time = time.time()
        self.current_student = student
        self.total_voltage = 0

    def update_voltage(self, voltage):
        if self.running:
            self.total_voltage += voltage
            self.data[self.cycle_id] = voltage

    def stop(self):
        """End session, log it into log.xlsx under sheet named YYYY-MM-DD, then reset."""
        if not self.running:
            return

        # --- Build the DataFrame for this session ---
        end_ts   = time.time()
        start_dt = datetime.fromtimestamp(self.start_time)
        end_dt   = datetime.fromtimestamp(end_ts)
        date_str = start_dt.strftime("%Y-%m-%d")  # sheet name

        df = pd.DataFrame([{
            "Student": self.current_student,
            "Cycle": self.cycle_id,
            "Start": start_dt.time(),
            "End":   end_dt.time(),
            "Duration (s)": end_dt.timestamp() - start_dt.timestamp(),
            "Total Voltage": self.total_voltage,
            "Energy (kWh)": self.total_voltage / 1000.0
        }])

        log_path = "log.xlsx"
        if os.path.exists(log_path):
            # load existing workbook just to get existing sheet info
            book = load_workbook(log_path)
            if date_str in book.sheetnames:
                ws = book[date_str]
                startrow = ws.max_row
                header   = False
            else:
                startrow = 0
                header   = True

            # append or create the date sheet
            with pd.ExcelWriter(
                log_path,
                engine="openpyxl",
                mode="a",
                if_sheet_exists="overlay",
                date_format="yyyy-mm-dd",
                datetime_format="hh:mm:ss"
            ) as writer:
                df.to_excel(
                    writer,
                    sheet_name=date_str,
                    index=False,
                    header=header,
                    startrow=startrow
                )
        else:
            # first run: make brand-new file
            with pd.ExcelWriter(
                log_path,
                engine="openpyxl",
                mode="w",
                date_format="yyyy-mm-dd",
                datetime_format="hh:mm:ss"
            ) as writer:
                df.to_excel(writer, sheet_name=date_str, index=False)

        # --- reset tracker state ---
        self.running = False
        self.current_student = None
        self.data = {}
        self.total_voltage = 0
